from __future__ import annotations

import os
import logging
import uuid
from datetime import datetime, timedelta, timezone, date
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Tuple

from dotenv import load_dotenv
from fastapi import APIRouter, Depends, FastAPI, HTTPException, Query, status
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import JWTError, jwt
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from pydantic import BaseModel, ConfigDict, EmailStr, Field
from starlette.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle
from reportlab.lib.units import cm

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

logger = logging.getLogger("cerdas_finansial")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)

# MongoDB connection (DO NOT hardcode; use env)
mongo_url = os.environ.get("MONGO_URL")
if not mongo_url:
    raise RuntimeError("MONGO_URL is not set")

client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get("DB_NAME", "test_database")]

# App + router
app = FastAPI(title="Cerdas Finansial API")
api_router = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer(auto_error=False)

JWT_SECRET = os.environ.get("JWT_SECRET", "dev")
JWT_ALG = "HS256"
JWT_EXPIRE_MINUTES = int(os.environ.get("JWT_EXPIRE_MINUTES", "10080"))

ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "moebarokocu@gmail.com").strip().lower()
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "261256")
ADMIN_NAME = os.environ.get("ADMIN_NAME", "Presiden Mubarak")


# -------------------------
# Utilities
# -------------------------

def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def gen_id() -> str:
    return str(uuid.uuid4())


def to_public(doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not doc:
        return None
    doc.pop("_id", None)
    return doc


def ym_from_str(month: str) -> Tuple[int, int]:
    # month: YYYY-MM
    try:
        y_str, m_str = month.split("-")
        y = int(y_str)
        m = int(m_str)
        if m < 1 or m > 12:
            raise ValueError
        return y, m
    except Exception as e:
        raise HTTPException(status_code=400, detail="month harus format YYYY-MM") from e


def date_from_str(d: str) -> date:
    try:
        return datetime.strptime(d, "%Y-%m-%d").date()
    except Exception as e:
        raise HTTPException(status_code=400, detail="date harus format YYYY-MM-DD") from e


def month_start_end(year: int, month: int) -> Tuple[date, date]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)
    return start, end


def rp(n: float) -> float:
    # IDR biasanya tanpa desimal. Untuk menghindari isu presisi float,
    # kita bulatkan ke rupiah terdekat.
    return float(int(round(float(n or 0), 0)))


# -------------------------
# Pydantic Schemas
# -------------------------

Role = Literal["user", "admin"]
Kind = Literal["income", "expense"]
TxType = Literal["income", "expense"]


class TokenResponse(BaseModel):
    token: str


class UserPublic(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str
    name: str
    email: EmailStr
    role: Role


class RegisterRequest(BaseModel):
    name: str = Field(min_length=2, max_length=80)
    email: EmailStr
    password: str = Field(min_length=4, max_length=128)
    confirm_password: str = Field(min_length=4, max_length=128)


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class PaymentMethodIn(BaseModel):
    name: str = Field(min_length=1, max_length=50)
    balance: float = 0


class PaymentMethodOut(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str
    name: str
    balance: float


class CategoryIn(BaseModel):
    kind: Kind
    name: str = Field(min_length=1, max_length=60)


class CategoryOut(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str
    kind: Kind
    name: str


class SubcategoryIn(BaseModel):
    kind: Kind
    category_id: str
    name: str = Field(min_length=1, max_length=60)


class SubcategoryOut(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str
    kind: Kind
    category_id: str
    name: str


class BudgetUpsertItem(BaseModel):
    subcategory_id: str
    amount: float


class BudgetBatchUpsert(BaseModel):
    month: str  # YYYY-MM
    items: List[BudgetUpsertItem]


class BudgetRow(BaseModel):
    subcategory_id: str
    subcategory_name: str
    category_name: str
    budget: float
    spent: float
    remaining: float
    percent: float


class BudgetOverviewResponse(BaseModel):
    month: str
    rows: List[BudgetRow]


class TransactionIn(BaseModel):
    type: TxType
    date: str  # YYYY-MM-DD
    category_id: str
    subcategory_id: str
    description: str = ""
    amount: float
    payment_method_id: str


class TransactionOut(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str
    type: TxType
    date: str
    category_id: str
    subcategory_id: str
    description: str
    amount: float
    payment_method_id: str
    created_at: str
    updated_at: str


class TransferIn(BaseModel):
    date: str  # YYYY-MM-DD
    from_payment_method_id: str
    to_payment_method_id: str
    amount: float
    description: str = ""


class TransferOut(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str
    date: str
    from_payment_method_id: str
    to_payment_method_id: str
    amount: float
    description: str
    created_at: str
    updated_at: str


class DailySpendPoint(BaseModel):
    date: str
    amount: float


class ReportExpenseRow(BaseModel):
    id: str
    date: str
    category_id: str
    category_name: str
    subcategory_id: str
    subcategory_name: str
    description: str
    amount: float
    payment_method_id: str
    payment_method_name: str


class ReportCategoryTotal(BaseModel):
    category_id: str
    category_name: str
    total: float


class ExpenseReportDataResponse(BaseModel):
    month: str
    total: float
    rows: List[ReportExpenseRow]
    totals_by_category: List[ReportCategoryTotal]


class DashboardOverviewResponse(BaseModel):
    month: str
    income_total: float
    expense_total: float
    net_total: float
    today_expense_total: float
    payment_methods: List[PaymentMethodOut]
    daily_expense: List[DailySpendPoint]
    budgets: List[BudgetRow]
    recent_transfers: List[TransferOut]


# -------------------------
# Auth helpers
# -------------------------


def hash_password(pw: str) -> str:
    return pwd_context.hash(pw)


def verify_password(pw: str, pw_hash: str) -> bool:
    return pwd_context.verify(pw, pw_hash)


def create_access_token(*, user_id: str, role: Role) -> str:
    exp = now_utc() + timedelta(minutes=JWT_EXPIRE_MINUTES)
    payload = {"sub": user_id, "role": role, "exp": exp}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def get_current_user(
    creds: Optional[HTTPAuthorizationCredentials] = Depends(security),
) -> Dict[str, Any]:
    if not creds or not creds.credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = creds.credentials
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError as e:
        raise HTTPException(status_code=401, detail="Invalid token") from e

    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User tidak ditemukan")
    return user


async def get_current_admin(user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return user


# -------------------------
# Seeding
# -------------------------

INCOME_SUBCATEGORIES = [
    "Gaji Bulanan",
    "Insentif",
    "Warisan Ayah",
    "Dari Orang Tua",
    "Hibah",
    "Khutbah",
    "Kajian",
    "Imam",
    "Barang Pribadi",
    "Barang Dagangan",
    "Refund",
]

DEFAULT_EXPENSE_CATEGORIES: Dict[str, List[str]] = {
    "Kebutuhan": ["Makan & Minum", "Transportasi", "Tagihan", "Bensin Motor", "Kesehatan"],
    "Keinginan": ["Hiburan", "Belanja", "Ngopi", "Travel"],
    "Investasi": ["Emas", "Reksadana", "Saham", "Crypto"],
    "Dana Darurat": ["Tabungan Darurat"],
}

DEFAULT_PAYMENT_METHODS = ["Cash", "GoPay", "Dana", "Bank"]


async def ensure_indexes() -> None:
    await db.users.create_index("email", unique=True)
    await db.payment_methods.create_index([("user_id", 1), ("name", 1)])
    await db.categories.create_index([("user_id", 1), ("kind", 1)])
    await db.subcategories.create_index([("user_id", 1), ("kind", 1), ("category_id", 1)])
    await db.budgets.create_index(
        [
            ("user_id", 1),
            ("year", 1),
            ("month", 1),
            ("subcategory_id", 1),
        ],
        unique=True,
    )
    await db.transactions.create_index([("user_id", 1), ("type", 1), ("date", 1)])
    await db.transfers.create_index([("user_id", 1), ("date", 1)])


async def seed_defaults_for_user(user_id: str) -> None:
    # Payment methods
    existing_pm = await db.payment_methods.count_documents({"user_id": user_id})
    if existing_pm == 0:
        docs = []
        for name in DEFAULT_PAYMENT_METHODS:
            docs.append(
                {
                    "id": gen_id(),
                    "user_id": user_id,
                    "name": name,
                    "balance": 0.0,
                    "created_at": now_utc().isoformat(),
                }
            )
        if docs:
            await db.payment_methods.insert_many(docs)

    # Categories + subcategories
    existing_cat = await db.categories.count_documents({"user_id": user_id})
    if existing_cat == 0:
        # Income category: Pemasukan
        income_cat_id = gen_id()
        await db.categories.insert_one(
            {
                "id": income_cat_id,
                "user_id": user_id,
                "kind": "income",
                "name": "Pemasukan",
                "created_at": now_utc().isoformat(),
            }
        )
        income_sub_docs = []
        for sc in INCOME_SUBCATEGORIES:
            income_sub_docs.append(
                {
                    "id": gen_id(),
                    "user_id": user_id,
                    "kind": "income",
                    "category_id": income_cat_id,
                    "name": sc,
                    "created_at": now_utc().isoformat(),
                }
            )
        if income_sub_docs:
            await db.subcategories.insert_many(income_sub_docs)

        # Expense categories
        expense_cat_ids: Dict[str, str] = {}
        exp_cat_docs = []
        for cat_name in DEFAULT_EXPENSE_CATEGORIES.keys():
            cid = gen_id()
            expense_cat_ids[cat_name] = cid
            exp_cat_docs.append(
                {
                    "id": cid,
                    "user_id": user_id,
                    "kind": "expense",
                    "name": cat_name,
                    "created_at": now_utc().isoformat(),
                }
            )
        if exp_cat_docs:
            await db.categories.insert_many(exp_cat_docs)

        exp_sub_docs = []
        for cat_name, subs in DEFAULT_EXPENSE_CATEGORIES.items():
            for sc in subs:
                exp_sub_docs.append(
                    {
                        "id": gen_id(),
                        "user_id": user_id,
                        "kind": "expense",
                        "category_id": expense_cat_ids[cat_name],
                        "name": sc,
                        "created_at": now_utc().isoformat(),
                    }
                )
        if exp_sub_docs:
            await db.subcategories.insert_many(exp_sub_docs)

    # Budgets for current month for expense subcategories
    today = datetime.now().date()
    year, month = today.year, today.month

    exp_subs = await db.subcategories.find({"user_id": user_id, "kind": "expense"}, {"_id": 0}).to_list(
        5000
    )
    budget_docs = []
    for sc in exp_subs:
        budget_docs.append(
            {
                "id": gen_id(),
                "user_id": user_id,
                "year": year,
                "month": month,
                "subcategory_id": sc["id"],
                "amount": 0.0,
            }
        )
    # upsert one by one to respect unique index
    for doc in budget_docs:
        await db.budgets.update_one(
            {
                "user_id": user_id,
                "year": year,
                "month": month,
                "subcategory_id": doc["subcategory_id"],
            },
            {"$setOnInsert": doc},
            upsert=True,
        )


async def seed_admin() -> None:
    admin = await db.users.find_one({"email": ADMIN_EMAIL}, {"_id": 0})
    if admin:
        if admin.get("role") != "admin":
            await db.users.update_one({"id": admin["id"]}, {"$set": {"role": "admin"}})
        return

    admin_doc = {
        "id": gen_id(),
        "name": ADMIN_NAME,
        "email": ADMIN_EMAIL,
        "password_hash": hash_password(ADMIN_PASSWORD),
        "role": "admin",
        "created_at": now_utc().isoformat(),
    }
    await db.users.insert_one(admin_doc)
    await seed_defaults_for_user(admin_doc["id"])
    logger.info("Seeded admin user: %s", ADMIN_EMAIL)


@app.on_event("startup")
async def on_startup() -> None:
    await ensure_indexes()
    await seed_admin()


@app.on_event("shutdown")
async def shutdown_db_client() -> None:
    client.close()



# -------------------------
# Reporting helpers
# -------------------------

MONTH_NAMES_ID = [
    "Januari",
    "Februari",
    "Maret",
    "April",
    "Mei",
    "Juni",
    "Juli",
    "Agustus",
    "September",
    "Oktober",
    "November",
    "Desember",
]


def safe_filename(s: str) -> str:
    keep = []
    for ch in s:
        if ch.isalnum() or ch in ["-", "_", ".", " "]:
            keep.append(ch)
    return "".join(keep).strip().replace(" ", "_")


async def build_expense_report_data(*, user: Dict[str, Any], month: str) -> ExpenseReportDataResponse:
    y, m = ym_from_str(month)
    start, end = month_start_end(y, m)

    txs = await db.transactions.find(
        {
            "user_id": user["id"],
            "type": "expense",
            "date": {"$gte": start.isoformat(), "$lt": end.isoformat()},
        },
        {
            "_id": 0,
            "id": 1,
            "date": 1,
            "category_id": 1,
            "subcategory_id": 1,
            "description": 1,
            "amount": 1,
            "payment_method_id": 1,
        },
    ).sort("date", 1).to_list(50000)

    cats = await db.categories.find(
        {"user_id": user["id"], "kind": "expense"},
        {"_id": 0, "id": 1, "name": 1},
    ).to_list(5000)
    subs = await db.subcategories.find(
        {"user_id": user["id"], "kind": "expense"},
        {"_id": 0, "id": 1, "name": 1},
    ).to_list(20000)
    pms = await db.payment_methods.find(
        {"user_id": user["id"]},
        {"_id": 0, "id": 1, "name": 1},
    ).to_list(2000)

    cat_by_id = {c["id"]: c["name"] for c in cats}
    sub_by_id = {s["id"]: s["name"] for s in subs}
    pm_by_id = {p["id"]: p["name"] for p in pms}

    rows: List[ReportExpenseRow] = []
    totals: Dict[str, float] = {}

    for tx in txs:
        cat_name = cat_by_id.get(tx.get("category_id"), "-")
        sub_name = sub_by_id.get(tx.get("subcategory_id"), "-")
        pm_name = pm_by_id.get(tx.get("payment_method_id"), "-")
        amt = rp(tx.get("amount", 0.0))
        cat_id = tx.get("category_id", "")
        totals[cat_id] = rp(totals.get(cat_id, 0.0) + amt)

        rows.append(
            ReportExpenseRow(
                id=tx["id"],
                date=tx.get("date", ""),
                category_id=cat_id,
                category_name=cat_name,
                subcategory_id=tx.get("subcategory_id", ""),
                subcategory_name=sub_name,
                description=tx.get("description", ""),
                amount=amt,
                payment_method_id=tx.get("payment_method_id", ""),
                payment_method_name=pm_name,
            )
        )

    totals_by_category = [
        ReportCategoryTotal(
            category_id=cid,
            category_name=cat_by_id.get(cid, "-"),
            total=rp(total),
        )
        for cid, total in totals.items()
        if cid
    ]
    totals_by_category.sort(key=lambda x: -x.total)

    total_all = rp(sum(r.amount for r in rows))

    return ExpenseReportDataResponse(
        month=month,
        total=total_all,
        rows=rows,
        totals_by_category=totals_by_category,
    )


def build_expense_pdf(*, user: Dict[str, Any], month: str, report: ExpenseReportDataResponse) -> bytes:
    y, m = ym_from_str(month)
    month_name = MONTH_NAMES_ID[m - 1]
    current_year = datetime.now().year

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title="Cerdas Finansial - Laporan Pengeluaran",
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = "Helvetica-Bold"
    title_style.fontSize = 18

    normal = styles["Normal"]
    normal.fontName = "Helvetica"
    normal.fontSize = 10

    elements: List[Any] = []

    elements.append(Paragraph("💵 CERDAS FINANSIAL", title_style))
    elements.append(Spacer(1, 6))
    elements.append(
        Paragraph(
            f"Laporan Pengeluaran Bulan <b>{month_name} {y}</b><br/>Pengguna: <b>{user.get('name','-')}</b>",
            normal,
        )
    )
    elements.append(Spacer(1, 10))

    # Table
    header = ["Tanggal", "Kategori", "Subkategori", "Deskripsi", "Metode", "Nominal (Rp)"]
    data: List[List[str]] = [header]

    for r in report.rows:
        data.append(
            [
                r.date,
                r.category_name,
                r.subcategory_name,
                (r.description or "-")[:60],
                r.payment_method_name,
                f"{int(r.amount):,}".replace(",", "."),
            ]
        )

    tbl = PdfTable(data, colWidths=[2.1 * cm, 2.5 * cm, 2.7 * cm, 5.5 * cm, 2.5 * cm, 2.6 * cm])
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FBC02D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DADFE3")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6FBFA")]),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
            ]
        )
    )

    elements.append(tbl)
    elements.append(Spacer(1, 10))

    # Totals by category
    if report.totals_by_category:
        elements.append(Paragraph("<b>Total per kategori</b>", normal))
        elements.append(Spacer(1, 6))
        sum_data = [["Kategori", "Total (Rp)"]]
        for t in report.totals_by_category:
            sum_data.append(
                [
                    t.category_name,
                    f"{int(t.total):,}".replace(",", "."),
                ]
            )
        sum_tbl = PdfTable(sum_data, colWidths=[10.3 * cm, 5.6 * cm])
        sum_tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9F7F5")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DADFE3")),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ]
            )
        )
        elements.append(sum_tbl)
        elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            f"<b>Total Pengeluaran Bulan Ini:</b> Rp {int(report.total):,}".replace(",", "."),
            normal,
        )
    )

    def on_page(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#6B7B7A"))
        canvas.drawRightString(
            A4[0] - 1.5 * cm,
            1.0 * cm,
            f"Dibuat otomatis oleh sistem Cerdas Finansial – © {current_year}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)
    return buf.getvalue()


def build_expense_xlsx_single_month(*, user: Dict[str, Any], month: str, report: ExpenseReportDataResponse) -> bytes:
    y, m = ym_from_str(month)
    month_name = MONTH_NAMES_ID[m - 1]
    sheet_name = f"Laporan_{month_name}_{y}"[:31]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    fill = PatternFill("solid", fgColor="FBC02D")
    thin = Side(style="thin", color="DADFE3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"].value = "CERDAS FINANSIAL"
    ws["A1"].font = title_font
    ws["A2"].value = f"Laporan Pengeluaran Bulan {month_name} {y}"
    ws["A3"].value = f"Pengguna: {user.get('name','-')}"

    start_row = 5
    columns = ["Tanggal", "Kategori", "Subkategori", "Deskripsi", "Nominal", "Metode Pembayaran"]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_cursor = start_row + 1
    for r in report.rows:
        values = [r.date, r.category_name, r.subcategory_name, r.description or "-", int(r.amount), r.payment_method_name]
        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=row_cursor, column=col_idx, value=v)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row_cursor, column=5).number_format = "#,##0"
        ws.cell(row=row_cursor, column=5).alignment = Alignment(horizontal="right", vertical="top")
        row_cursor += 1

    row_cursor += 1
    ws.cell(row=row_cursor, column=4, value="Total Pengeluaran Bulan Ini:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=5, value=int(report.total)).font = Font(bold=True)
    ws.cell(row=row_cursor, column=5).number_format = "#,##0"
    ws.cell(row=row_cursor, column=5).alignment = Alignment(horizontal="right")

    # Totals by category section
    row_cursor += 2
    ws.cell(row=row_cursor, column=1, value="Total per Kategori").font = Font(bold=True)
    row_cursor += 1
    ws.cell(row=row_cursor, column=1, value="Kategori").font = header_font
    ws.cell(row=row_cursor, column=2, value="Total").font = header_font
    ws.cell(row=row_cursor, column=1).fill = PatternFill("solid", fgColor="E9F7F5")
    ws.cell(row=row_cursor, column=2).fill = PatternFill("solid", fgColor="E9F7F5")
    row_cursor += 1
    for t in report.totals_by_category:
        ws.cell(row=row_cursor, column=1, value=t.category_name)
        ws.cell(row=row_cursor, column=2, value=int(t.total)).number_format = "#,##0"
        row_cursor += 1

    # Auto width
    for col in range(1, 7):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 50)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


async def build_expense_xlsx_year(*, user: Dict[str, Any], year: int) -> bytes:
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    for m in range(1, 13):
        month = f"{year}-{str(m).zfill(2)}"
        report = await build_expense_report_data(user=user, month=month)
        month_name = MONTH_NAMES_ID[m - 1]
        ws = wb.create_sheet(title=f"{month_name}_{year}"[:31])

        ws.append(["Tanggal", "Kategori", "Subkategori", "Deskripsi", "Nominal", "Metode Pembayaran"])
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="FBC02D")
            c.alignment = Alignment(horizontal="center")

        for r in report.rows:
            ws.append([r.date, r.category_name, r.subcategory_name, r.description or "-", int(r.amount), r.payment_method_name])
        ws.append([])
        ws.append(["", "", "", "Total Pengeluaran Bulan Ini:", int(report.total), ""])
        ws[ws.max_row][5].number_format = "#,##0"

        # auto width
        for col in range(1, 7):
            max_len = 0
            for cell in ws[get_column_letter(col)]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 50)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# -------------------------
# Core accounting helpers
# -------------------------


async def adjust_payment_method_balance(*, user_id: str, payment_method_id: str, delta: float) -> None:
    res = await db.payment_methods.update_one(
        {"id": payment_method_id, "user_id": user_id},
        {"$inc": {"balance": rp(delta)}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Metode pembayaran tidak ditemukan")


async def apply_transaction_effect(user_id: str, tx: Dict[str, Any], direction: Literal["apply", "revert"]) -> None:
    # tx.amount: positive
    amt = float(tx["amount"])
    if tx["type"] == "income":
        delta = amt
    else:
        delta = -amt

    if direction == "revert":
        delta = -delta

    await adjust_payment_method_balance(user_id=user_id, payment_method_id=tx["payment_method_id"], delta=delta)


async def apply_transfer_effect(user_id: str, tr: Dict[str, Any], direction: Literal["apply", "revert"]) -> None:
    amt = float(tr["amount"])
    from_delta = -amt
    to_delta = amt
    if direction == "revert":
        from_delta = -from_delta
        to_delta = -to_delta

    await adjust_payment_method_balance(user_id=user_id, payment_method_id=tr["from_payment_method_id"], delta=from_delta)
    await adjust_payment_method_balance(user_id=user_id, payment_method_id=tr["to_payment_method_id"], delta=to_delta)


# -------------------------
# Routes
# -------------------------


@api_router.get("/")
async def root() -> Dict[str, str]:
    return {"message": "Cerdas Finansial API"}


# ---- Auth ----


@api_router.post("/auth/register", response_model=TokenResponse)
async def register(payload: RegisterRequest) -> TokenResponse:
    if payload.password != payload.confirm_password:
        raise HTTPException(status_code=400, detail="Konfirmasi password tidak sama")

    email = payload.email.strip().lower()
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")

    user_doc = {
        "id": gen_id(),
        "name": payload.name.strip(),
        "email": email,
        "password_hash": hash_password(payload.password),
        "role": "user",
        "created_at": now_utc().isoformat(),
    }
    await db.users.insert_one(user_doc)
    await seed_defaults_for_user(user_doc["id"])
    token = create_access_token(user_id=user_doc["id"], role="user")
    return TokenResponse(token=token)


@api_router.post("/auth/login", response_model=TokenResponse)
async def login(payload: LoginRequest) -> TokenResponse:
    email = payload.email.strip().lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=400, detail="Email atau password salah")
    if not verify_password(payload.password, user.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Email atau password salah")

    token = create_access_token(user_id=user["id"], role=user.get("role", "user"))
    return TokenResponse(token=token)


@api_router.get("/auth/me", response_model=UserPublic)
async def me(user: Dict[str, Any] = Depends(get_current_user)) -> UserPublic:
    return UserPublic(id=user["id"], name=user["name"], email=user["email"], role=user.get("role", "user"))


# ---- Payment Methods ----


@api_router.get("/payment-methods", response_model=List[PaymentMethodOut])
async def list_payment_methods(user: Dict[str, Any] = Depends(get_current_user)) -> List[PaymentMethodOut]:
    methods = await db.payment_methods.find({"user_id": user["id"]}, {"_id": 0}).sort("name", 1).to_list(1000)
    return methods


@api_router.post("/payment-methods", response_model=PaymentMethodOut)
async def create_payment_method(
    payload: PaymentMethodIn, user: Dict[str, Any] = Depends(get_current_user)
) -> PaymentMethodOut:
    doc = {
        "id": gen_id(),
        "user_id": user["id"],
        "name": payload.name.strip(),
        "balance": rp(payload.balance),
        "created_at": now_utc().isoformat(),
    }
    await db.payment_methods.insert_one(doc)
    return doc


@api_router.put("/payment-methods/{payment_method_id}", response_model=PaymentMethodOut)
async def update_payment_method(
    payment_method_id: str, payload: PaymentMethodIn, user: Dict[str, Any] = Depends(get_current_user)
) -> PaymentMethodOut:
    await db.payment_methods.update_one(
        {"id": payment_method_id, "user_id": user["id"]},
        {"$set": {"name": payload.name.strip(), "balance": rp(payload.balance)}},
    )
    updated = await db.payment_methods.find_one({"id": payment_method_id, "user_id": user["id"]}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Metode pembayaran tidak ditemukan")
    return updated


@api_router.delete("/payment-methods/{payment_method_id}")
async def delete_payment_method(payment_method_id: str, user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, str]:
    # Guard: cannot delete if referenced by transactions/transfers
    tx_count = await db.transactions.count_documents({"user_id": user["id"], "payment_method_id": payment_method_id})
    tr_count = await db.transfers.count_documents(
        {
            "user_id": user["id"],
            "$or": [
                {"from_payment_method_id": payment_method_id},
                {"to_payment_method_id": payment_method_id},
            ],
        }
    )
    if tx_count > 0 or tr_count > 0:
        raise HTTPException(status_code=400, detail="Tidak bisa hapus: metode masih dipakai transaksi/transfer")

    res = await db.payment_methods.delete_one({"id": payment_method_id, "user_id": user["id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Metode pembayaran tidak ditemukan")
    return {"status": "ok"}


# ---- Categories ----


@api_router.get("/categories", response_model=List[CategoryOut])
async def list_categories(
    kind: Kind = Query(...),
    user: Dict[str, Any] = Depends(get_current_user),
) -> List[CategoryOut]:
    cats = await db.categories.find({"user_id": user["id"], "kind": kind}, {"_id": 0}).sort("name", 1).to_list(1000)
    return cats


@api_router.post("/categories", response_model=CategoryOut)
async def create_category(payload: CategoryIn, user: Dict[str, Any] = Depends(get_current_user)) -> CategoryOut:
    doc = {
        "id": gen_id(),
        "user_id": user["id"],
        "kind": payload.kind,
        "name": payload.name.strip(),
        "created_at": now_utc().isoformat(),
    }
    await db.categories.insert_one(doc)
    return doc


@api_router.put("/categories/{category_id}", response_model=CategoryOut)
async def update_category(category_id: str, payload: CategoryIn, user: Dict[str, Any] = Depends(get_current_user)) -> CategoryOut:
    await db.categories.update_one(
        {"id": category_id, "user_id": user["id"]},
        {"$set": {"name": payload.name.strip()}},
    )
    updated = await db.categories.find_one({"id": category_id, "user_id": user["id"]}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Kategori tidak ditemukan")
    return updated


@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, str]:
    sub_count = await db.subcategories.count_documents({"user_id": user["id"], "category_id": category_id})
    if sub_count > 0:
        raise HTTPException(status_code=400, detail="Hapus subkategori dulu sebelum hapus kategori")

    res = await db.categories.delete_one({"id": category_id, "user_id": user["id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kategori tidak ditemukan")
    return {"status": "ok"}


# ---- Subcategories ----


@api_router.get("/subcategories", response_model=List[SubcategoryOut])
async def list_subcategories(
    kind: Kind = Query(...),
    category_id: Optional[str] = Query(None),
    user: Dict[str, Any] = Depends(get_current_user),
) -> List[SubcategoryOut]:
    q: Dict[str, Any] = {"user_id": user["id"], "kind": kind}
    if category_id:
        q["category_id"] = category_id
    subs = await db.subcategories.find(q, {"_id": 0}).sort("name", 1).to_list(5000)
    return subs


@api_router.post("/subcategories", response_model=SubcategoryOut)
async def create_subcategory(payload: SubcategoryIn, user: Dict[str, Any] = Depends(get_current_user)) -> SubcategoryOut:
    cat = await db.categories.find_one({"id": payload.category_id, "user_id": user["id"], "kind": payload.kind})
    if not cat:
        raise HTTPException(status_code=400, detail="Kategori tidak valid")

    doc = {
        "id": gen_id(),
        "user_id": user["id"],
        "kind": payload.kind,
        "category_id": payload.category_id,
        "name": payload.name.strip(),
        "created_at": now_utc().isoformat(),
    }
    await db.subcategories.insert_one(doc)

    # ensure budget row exists for this expense subcategory for current month
    if payload.kind == "expense":
        today = datetime.now().date()
        await db.budgets.update_one(
            {
                "user_id": user["id"],
                "year": today.year,
                "month": today.month,
                "subcategory_id": doc["id"],
            },
            {
                "$setOnInsert": {
                    "id": gen_id(),
                    "user_id": user["id"],
                    "year": today.year,
                    "month": today.month,
                    "subcategory_id": doc["id"],
                    "amount": 0.0,
                }
            },
            upsert=True,
        )

    return doc


@api_router.put("/subcategories/{subcategory_id}", response_model=SubcategoryOut)
async def update_subcategory(subcategory_id: str, payload: SubcategoryIn, user: Dict[str, Any] = Depends(get_current_user)) -> SubcategoryOut:
    cat = await db.categories.find_one({"id": payload.category_id, "user_id": user["id"], "kind": payload.kind})
    if not cat:
        raise HTTPException(status_code=400, detail="Kategori tidak valid")

    await db.subcategories.update_one(
        {"id": subcategory_id, "user_id": user["id"]},
        {"$set": {"name": payload.name.strip(), "category_id": payload.category_id, "kind": payload.kind}},
    )
    updated = await db.subcategories.find_one({"id": subcategory_id, "user_id": user["id"]}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Subkategori tidak ditemukan")
    return updated


@api_router.delete("/subcategories/{subcategory_id}")
async def delete_subcategory(subcategory_id: str, user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, str]:
    tx_count = await db.transactions.count_documents({"user_id": user["id"], "subcategory_id": subcategory_id})
    if tx_count > 0:
        raise HTTPException(status_code=400, detail="Tidak bisa hapus: subkategori dipakai transaksi")

    await db.budgets.delete_many({"user_id": user["id"], "subcategory_id": subcategory_id})
    res = await db.subcategories.delete_one({"id": subcategory_id, "user_id": user["id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Subkategori tidak ditemukan")
    return {"status": "ok"}


# ---- Budgets ----


@api_router.get("/budgets/overview", response_model=BudgetOverviewResponse)
async def budgets_overview(
    month: str = Query(...),
    user: Dict[str, Any] = Depends(get_current_user),
) -> BudgetOverviewResponse:
    y, m = ym_from_str(month)
    start, end = month_start_end(y, m)

    subs = await db.subcategories.find(
        {"user_id": user["id"], "kind": "expense"},
        {"_id": 0, "id": 1, "name": 1, "category_id": 1},
    ).to_list(5000)
    cats = await db.categories.find(
        {"user_id": user["id"], "kind": "expense"},
        {"_id": 0, "id": 1, "name": 1},
    ).to_list(1000)
    cat_by_id = {c["id"]: c for c in cats}

    budgets = await db.budgets.find(
        {"user_id": user["id"], "year": y, "month": m},
        {"_id": 0, "subcategory_id": 1, "amount": 1},
    ).to_list(5000)
    budget_by_sc = {b["subcategory_id"]: float(b.get("amount", 0.0)) for b in budgets}

    # spending per subcategory for month
    expense_txs = await db.transactions.find(
        {
            "user_id": user["id"],
            "type": "expense",
            "date": {"$gte": start.isoformat(), "$lt": end.isoformat()},
        },
        {"_id": 0, "subcategory_id": 1, "amount": 1},
    ).to_list(10000)
    spent_by_sc: Dict[str, float] = {}
    for tx in expense_txs:
        scid = tx["subcategory_id"]
        spent_by_sc[scid] = float(spent_by_sc.get(scid, 0.0)) + float(tx.get("amount", 0.0))

    rows: List[BudgetRow] = []
    for sc in subs:
        budget = float(budget_by_sc.get(sc["id"], 0.0))
        spent = float(spent_by_sc.get(sc["id"], 0.0))
        remaining = budget - spent
        percent = 0.0
        if budget > 0:
            percent = max(0.0, min(100.0, (spent / budget) * 100.0))
        cat_name = cat_by_id.get(sc["category_id"], {}).get("name", "-")
        rows.append(
            BudgetRow(
                subcategory_id=sc["id"],
                subcategory_name=sc["name"],
                category_name=cat_name,
                budget=rp(budget),
                spent=rp(spent),
                remaining=rp(remaining),
                percent=rp(percent),
            )
        )

    # sort by percent desc then name
    rows.sort(key=lambda r: (-r.percent, r.subcategory_name.lower()))

    return BudgetOverviewResponse(month=month, rows=rows)


@api_router.put("/budgets")
async def budgets_upsert(payload: BudgetBatchUpsert, user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, str]:
    y, m = ym_from_str(payload.month)

    for item in payload.items:
        # validate subcategory belongs to user and is expense
        sc = await db.subcategories.find_one(
            {"id": item.subcategory_id, "user_id": user["id"], "kind": "expense"},
            {"_id": 0},
        )
        if not sc:
            raise HTTPException(status_code=400, detail="Subkategori tidak valid")

        await db.budgets.update_one(
            {
                "user_id": user["id"],
                "year": y,
                "month": m,
                "subcategory_id": item.subcategory_id,
            },
            {
                "$set": {"amount": rp(item.amount)},
                "$setOnInsert": {
                    "id": gen_id(),
                    "user_id": user["id"],
                    "year": y,
                    "month": m,
                    "subcategory_id": item.subcategory_id,
                },
            },
            upsert=True,
        )

    return {"status": "ok"}


# ---- Reports (Expense Export) ----


@api_router.get("/reports/expenses/data", response_model=ExpenseReportDataResponse)
async def expense_report_data(
    month: str = Query(...),
    user: Dict[str, Any] = Depends(get_current_user),
) -> ExpenseReportDataResponse:
    return await build_expense_report_data(user=user, month=month)


@api_router.get("/reports/expenses/pdf")
async def expense_report_pdf(
    month: str = Query(...),
    user: Dict[str, Any] = Depends(get_current_user),
):
    report = await build_expense_report_data(user=user, month=month)
    pdf_bytes = build_expense_pdf(user=user, month=month, report=report)

    filename = safe_filename(f"Laporan_Pengeluaran_{month}_{user.get('name','User')}.pdf")
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
        "Cache-Control": "no-store",
    }
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)


@api_router.get("/reports/expenses/xlsx")
async def expense_report_xlsx(
    month: str = Query(...),
    user: Dict[str, Any] = Depends(get_current_user),
):
    report = await build_expense_report_data(user=user, month=month)
    xlsx_bytes = build_expense_xlsx_single_month(user=user, month=month, report=report)

    filename = safe_filename(f"Laporan_Pengeluaran_{month}_{user.get('name','User')}.xlsx")
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@api_router.get("/reports/expenses/xlsx-year")
async def expense_report_xlsx_year(
    year: int = Query(..., ge=2000, le=2100),
    user: Dict[str, Any] = Depends(get_current_user),
):
    xlsx_bytes = await build_expense_xlsx_year(user=user, year=year)

    filename = safe_filename(f"Laporan_Pengeluaran_{year}_{user.get('name','User')}.xlsx")
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ---- Transactions ----


@api_router.get("/transactions", response_model=List[TransactionOut])
async def list_transactions(
    type: TxType = Query(...),
    month: str = Query(...),
    user: Dict[str, Any] = Depends(get_current_user),
) -> List[TransactionOut]:
    y, m = ym_from_str(month)
    start, end = month_start_end(y, m)
    txs = await db.transactions.find(
        {
            "user_id": user["id"],
            "type": type,
            "date": {"$gte": start.isoformat(), "$lt": end.isoformat()},
        },
        {"_id": 0},
    ).sort("date", -1).to_list(10000)
    return txs


@api_router.post("/transactions", response_model=TransactionOut)
async def create_transaction(payload: TransactionIn, user: Dict[str, Any] = Depends(get_current_user)) -> TransactionOut:
    _ = date_from_str(payload.date)

    # validate references belong to user
    cat = await db.categories.find_one({"id": payload.category_id, "user_id": user["id"], "kind": payload.type}, {"_id": 0})
    if not cat:
        raise HTTPException(status_code=400, detail="Kategori tidak valid")
    sc = await db.subcategories.find_one(
        {
            "id": payload.subcategory_id,
            "user_id": user["id"],
            "kind": payload.type,
            "category_id": payload.category_id,
        },
        {"_id": 0},
    )
    if not sc:
        raise HTTPException(status_code=400, detail="Subkategori tidak valid")

    pm = await db.payment_methods.find_one({"id": payload.payment_method_id, "user_id": user["id"]}, {"_id": 0})
    if not pm:
        raise HTTPException(status_code=400, detail="Metode pembayaran tidak valid")

    doc = {
        "id": gen_id(),
        "user_id": user["id"],
        "type": payload.type,
        "date": payload.date,
        "category_id": payload.category_id,
        "subcategory_id": payload.subcategory_id,
        "description": (payload.description or "").strip(),
        "amount": rp(payload.amount),
        "payment_method_id": payload.payment_method_id,
        "created_at": now_utc().isoformat(),
        "updated_at": now_utc().isoformat(),
    }

    await db.transactions.insert_one(doc)
    await apply_transaction_effect(user["id"], doc, "apply")
    return doc


@api_router.put("/transactions/{transaction_id}", response_model=TransactionOut)
async def update_transaction(
    transaction_id: str, payload: TransactionIn, user: Dict[str, Any] = Depends(get_current_user)
) -> TransactionOut:
    _ = date_from_str(payload.date)

    existing = await db.transactions.find_one({"id": transaction_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")

    # validate
    cat = await db.categories.find_one({"id": payload.category_id, "user_id": user["id"], "kind": payload.type}, {"_id": 0})
    if not cat:
        raise HTTPException(status_code=400, detail="Kategori tidak valid")

    sc = await db.subcategories.find_one(
        {
            "id": payload.subcategory_id,
            "user_id": user["id"],
            "kind": payload.type,
            "category_id": payload.category_id,
        },
        {"_id": 0},
    )
    if not sc:
        raise HTTPException(status_code=400, detail="Subkategori tidak valid")

    pm = await db.payment_methods.find_one({"id": payload.payment_method_id, "user_id": user["id"]}, {"_id": 0})
    if not pm:
        raise HTTPException(status_code=400, detail="Metode pembayaran tidak valid")

    # revert old effect then apply new
    await apply_transaction_effect(user["id"], existing, "revert")

    new_doc = {
        **existing,
        "type": payload.type,
        "date": payload.date,
        "category_id": payload.category_id,
        "subcategory_id": payload.subcategory_id,
        "description": (payload.description or "").strip(),
        "amount": rp(payload.amount),
        "payment_method_id": payload.payment_method_id,
        "updated_at": now_utc().isoformat(),
    }

    await db.transactions.update_one(
        {"id": transaction_id, "user_id": user["id"]},
        {"$set": {k: v for k, v in new_doc.items() if k != "id"}},
    )

    await apply_transaction_effect(user["id"], new_doc, "apply")

    updated = await db.transactions.find_one({"id": transaction_id, "user_id": user["id"]}, {"_id": 0})
    return updated


@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str, user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, str]:
    existing = await db.transactions.find_one({"id": transaction_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")

    await apply_transaction_effect(user["id"], existing, "revert")
    await db.transactions.delete_one({"id": transaction_id, "user_id": user["id"]})
    return {"status": "ok"}


# ---- Transfers ----


@api_router.get("/transfers", response_model=List[TransferOut])
async def list_transfers(
    month: str = Query(...),
    user: Dict[str, Any] = Depends(get_current_user),
) -> List[TransferOut]:
    y, m = ym_from_str(month)
    start, end = month_start_end(y, m)
    trs = await db.transfers.find(
        {"user_id": user["id"], "date": {"$gte": start.isoformat(), "$lt": end.isoformat()}},
        {"_id": 0},
    ).sort("date", -1).to_list(10000)
    return trs


@api_router.post("/transfers", response_model=TransferOut)
async def create_transfer(payload: TransferIn, user: Dict[str, Any] = Depends(get_current_user)) -> TransferOut:
    _ = date_from_str(payload.date)

    if payload.from_payment_method_id == payload.to_payment_method_id:
        raise HTTPException(status_code=400, detail="Metode asal dan tujuan tidak boleh sama")

    # validate payment methods
    from_pm = await db.payment_methods.find_one({"id": payload.from_payment_method_id, "user_id": user["id"]}, {"_id": 0})
    to_pm = await db.payment_methods.find_one({"id": payload.to_payment_method_id, "user_id": user["id"]}, {"_id": 0})
    if not from_pm or not to_pm:
        raise HTTPException(status_code=400, detail="Metode pembayaran tidak valid")

    doc = {
        "id": gen_id(),
        "user_id": user["id"],
        "date": payload.date,
        "from_payment_method_id": payload.from_payment_method_id,
        "to_payment_method_id": payload.to_payment_method_id,
        "amount": rp(payload.amount),
        "description": (payload.description or "").strip(),
        "created_at": now_utc().isoformat(),
        "updated_at": now_utc().isoformat(),
    }

    await db.transfers.insert_one(doc)
    await apply_transfer_effect(user["id"], doc, "apply")
    return doc


@api_router.put("/transfers/{transfer_id}", response_model=TransferOut)
async def update_transfer(transfer_id: str, payload: TransferIn, user: Dict[str, Any] = Depends(get_current_user)) -> TransferOut:
    _ = date_from_str(payload.date)

    if payload.from_payment_method_id == payload.to_payment_method_id:
        raise HTTPException(status_code=400, detail="Metode asal dan tujuan tidak boleh sama")

    existing = await db.transfers.find_one({"id": transfer_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Transfer tidak ditemukan")

    from_pm = await db.payment_methods.find_one({"id": payload.from_payment_method_id, "user_id": user["id"]}, {"_id": 0})
    to_pm = await db.payment_methods.find_one({"id": payload.to_payment_method_id, "user_id": user["id"]}, {"_id": 0})
    if not from_pm or not to_pm:
        raise HTTPException(status_code=400, detail="Metode pembayaran tidak valid")

    await apply_transfer_effect(user["id"], existing, "revert")

    new_doc = {
        **existing,
        "date": payload.date,
        "from_payment_method_id": payload.from_payment_method_id,
        "to_payment_method_id": payload.to_payment_method_id,
        "amount": rp(payload.amount),
        "description": (payload.description or "").strip(),
        "updated_at": now_utc().isoformat(),
    }

    await db.transfers.update_one(
        {"id": transfer_id, "user_id": user["id"]},
        {"$set": {k: v for k, v in new_doc.items() if k != "id"}},
    )

    await apply_transfer_effect(user["id"], new_doc, "apply")

    updated = await db.transfers.find_one({"id": transfer_id, "user_id": user["id"]}, {"_id": 0})
    return updated


@api_router.delete("/transfers/{transfer_id}")
async def delete_transfer(transfer_id: str, user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, str]:
    existing = await db.transfers.find_one({"id": transfer_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Transfer tidak ditemukan")

    await apply_transfer_effect(user["id"], existing, "revert")
    await db.transfers.delete_one({"id": transfer_id, "user_id": user["id"]})
    return {"status": "ok"}


# ---- Dashboard ----


@api_router.get("/dashboard/overview", response_model=DashboardOverviewResponse)
async def dashboard_overview(
    month: str = Query(...),
    days: int = Query(30, ge=7, le=90),
    user: Dict[str, Any] = Depends(get_current_user),
) -> DashboardOverviewResponse:
    y, m = ym_from_str(month)
    start, end = month_start_end(y, m)

    income_txs = await db.transactions.find(
        {"user_id": user["id"], "type": "income", "date": {"$gte": start.isoformat(), "$lt": end.isoformat()}},
        {"_id": 0, "amount": 1, "date": 1},
    ).to_list(20000)
    expense_txs = await db.transactions.find(
        {"user_id": user["id"], "type": "expense", "date": {"$gte": start.isoformat(), "$lt": end.isoformat()}},
        {"_id": 0, "amount": 1, "date": 1},
    ).to_list(20000)

    income_total = sum(float(t.get("amount", 0.0)) for t in income_txs)
    expense_total = sum(float(t.get("amount", 0.0)) for t in expense_txs)
    net_total = income_total - expense_total

    today_iso = datetime.now().date().isoformat()
    today_expense_total = sum(float(t.get("amount", 0.0)) for t in expense_txs if t.get("date") == today_iso)

    methods = await db.payment_methods.find({"user_id": user["id"]}, {"_id": 0}).sort("name", 1).to_list(1000)

    # daily expense last N days (ending today) within selected month (if month not current, still compute within that month)
    # We'll compute points from max(start, end-days) to min(end-1, today) if same month.
    end_date = min(end, datetime.now().date() + timedelta(days=1))
    start_days = end_date - timedelta(days=days)
    start_date = max(start, start_days)

    # group expense by date
    daily_map: Dict[str, float] = {}
    for t in expense_txs:
        d = t.get("date")
        if not d:
            continue
        if d < start_date.isoformat() or d >= end_date.isoformat():
            continue
        daily_map[d] = daily_map.get(d, 0.0) + float(t.get("amount", 0.0))

    points: List[DailySpendPoint] = []
    cursor = start_date
    while cursor < end_date:
        iso = cursor.isoformat()
        points.append(DailySpendPoint(date=iso, amount=rp(daily_map.get(iso, 0.0))))
        cursor = cursor + timedelta(days=1)

    budget_resp = await budgets_overview(month=month, user=user)  # reuse logic

    transfers = await db.transfers.find(
        {"user_id": user["id"], "date": {"$gte": start.isoformat(), "$lt": end.isoformat()}},
        {"_id": 0},
    ).sort("date", -1).limit(20).to_list(20)

    return DashboardOverviewResponse(
        month=month,
        income_total=rp(income_total),
        expense_total=rp(expense_total),
        net_total=rp(net_total),
        today_expense_total=rp(today_expense_total),
        payment_methods=methods,
        daily_expense=points,
        budgets=budget_resp.rows,
        recent_transfers=transfers,
    )


# ---- Admin ----


@api_router.get("/admin/users", response_model=List[UserPublic])
async def admin_list_users(admin: Dict[str, Any] = Depends(get_current_admin)) -> List[UserPublic]:
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(5000)
    return users


@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, admin: Dict[str, Any] = Depends(get_current_admin)) -> Dict[str, str]:
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="Tidak bisa menghapus akun admin sendiri")

    await db.transactions.delete_many({"user_id": user_id})
    await db.transfers.delete_many({"user_id": user_id})
    await db.budgets.delete_many({"user_id": user_id})
    await db.subcategories.delete_many({"user_id": user_id})
    await db.categories.delete_many({"user_id": user_id})
    await db.payment_methods.delete_many({"user_id": user_id})
    res = await db.users.delete_one({"id": user_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    return {"status": "ok"}


# Include router
app.include_router(api_router)
